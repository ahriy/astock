"""
数据导出模块 - 将分析结果导出为Excel、CSV等格式
包含格式化和美化功能
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Optional
from datetime import datetime
import os
import logging
from config import OUTPUT_DIR

logger = logging.getLogger(__name__)


class DataExporter:
    """数据导出器类"""
    
    def __init__(self, output_dir: str = OUTPUT_DIR):
        """
        初始化导出器
        
        Args:
            output_dir: 输出目录路径
        """
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        logger.info(f"DataExporter 初始化完成，输出目录: {output_dir}")
    
    def export_to_excel(self, 
                       df: pd.DataFrame, 
                       filename: str = None,
                       sheet_name: str = "分析结果",
                       add_formatting: bool = True) -> str:
        """
        导出为Excel文件
        
        Args:
            df: 要导出的DataFrame
            filename: 文件名（不含扩展名），默认使用时间戳
            sheet_name: 工作表名称
            add_formatting: 是否添加格式
            
        Returns:
            导出的文件路径
        """
        if filename is None:
            filename = f"a_stock_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        filepath = os.path.join(self.output_dir, f"{filename}.xlsx")
        
        # 创建Excel writer
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            if add_formatting:
                self._format_excel(writer, df, sheet_name)
        
        logger.info(f"Excel文件已导出: {filepath}")
        return filepath
    
    def _format_excel(self, writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str):
        """格式化Excel文件"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # 定义样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # 设置列宽
            for i, column in enumerate(df.columns):
                max_length = max(
                    df[column].astype(str).map(len).max() if len(df) > 0 else 0,
                    len(str(column))
                )
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[get_column_letter(i + 1)].width = adjusted_width
            
            # 格式化表头
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # 冻结首行
            worksheet.freeze_panes = "A2"
            
        except Exception as e:
            logger.warning(f"Excel格式化失败: {e}")
    
    def export_to_csv(self, 
                     df: pd.DataFrame, 
                     filename: str = None,
                     encoding: str = 'utf-8-sig') -> str:
        """
        导出为CSV文件
        
        Args:
            df: 要导出的DataFrame
            filename: 文件名（不含扩展名），默认使用时间戳
            encoding: 编码格式
            
        Returns:
            导出的文件路径
        """
        if filename is None:
            filename = f"a_stock_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        filepath = os.path.join(self.output_dir, f"{filename}.csv")
        df.to_csv(filepath, index=False, encoding=encoding)
        
        logger.info(f"CSV文件已导出: {filepath}")
        return filepath
    
    def export_by_industry(self, 
                          df: pd.DataFrame,
                          base_filename: str = None) -> List[str]:
        """
        按行业分别导出Excel文件
        
        Args:
            df: 分析结果DataFrame
            base_filename: 基础文件名
            
        Returns:
            导出的文件路径列表
        """
        if base_filename is None:
            base_filename = f"a_stock_by_industry_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        exported_files = []
        
        if "所属行业" not in df.columns:
            logger.warning("数据中缺少'所属行业'字段，无法按行业导出")
            return exported_files
        
        # 按行业分组
        industries = df["所属行业"].unique()
        
        for industry in industries:
            if pd.isna(industry) or industry == "":
                continue
            
            industry_df = df[df["所属行业"] == industry]
            
            # 清理文件名
            safe_industry = str(industry).replace("/", "_").replace("\\", "_")
            filename = f"{base_filename}_{safe_industry}"
            filepath = os.path.join(self.output_dir, f"{filename}.xlsx")
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                industry_df.to_excel(writer, sheet_name=safe_industry[:31], index=False)
            
            exported_files.append(filepath)
            logger.info(f"行业 '{industry}' 导出完成: {filepath}")
        
        return exported_files
    
    def export_top_stocks(self,
                         df: pd.DataFrame,
                         top_n: int = 50,
                         sort_by: str = "价值评分",
                         filename: str = None) -> str:
        """
        导出排名靠前的股票
        
        Args:
            df: 分析结果DataFrame
            top_n: 取前N名
            sort_by: 排序字段
            filename: 文件名
            
        Returns:
            导出的文件路径
        """
        if filename is None:
            filename = f"top_{top_n}_stocks_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        # 排序并取前N
        if sort_by in df.columns:
            top_df = df.sort_values(sort_by, ascending=False).head(top_n)
        else:
            top_df = df.head(top_n)
        
        return self.export_to_excel(top_df, filename, f"Top{top_n}")
    
    def export_comprehensive_report(self,
                                   df: pd.DataFrame,
                                   summary_stats: Dict = None,
                                   filename: str = None) -> str:
        """
        导出综合报告（包含多个工作表）
        
        Args:
            df: 分析结果DataFrame
            summary_stats: 汇总统计信息
            filename: 文件名
            
        Returns:
            导出的文件路径
        """
        if filename is None:
            filename = f"comprehensive_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        filepath = os.path.join(self.output_dir, f"{filename}.xlsx")
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # 工作表1: 全部数据
            df.to_excel(writer, sheet_name="全部股票", index=False)
            self._format_excel(writer, df, "全部股票")
            
            # 工作表2: 按评级分类
            if "评级" in df.columns:
                rating_summary = df.groupby("评级").agg({
                    "股票代码": "count",
                    "价值评分": "mean",
                    "PE_3年均值": "mean",
                    "PB_3年均值": "mean",
                    "股息率_3年均值": "mean"
                }).round(2)
                rating_summary.columns = ["股票数量", "平均评分", "平均PE", "平均PB", "平均股息率"]
                rating_summary.to_excel(writer, sheet_name="评级分布")
            
            # 工作表3: 行业统计
            if "所属行业" in df.columns:
                industry_summary = df.groupby("所属行业").agg({
                    "股票代码": "count",
                    "价值评分": "mean",
                    "PE_3年均值": "mean",
                    "PB_3年均值": "mean"
                }).round(2)
                industry_summary.columns = ["股票数量", "平均评分", "平均PE", "平均PB"]
                industry_summary = industry_summary.sort_values("股票数量", ascending=False)
                industry_summary.to_excel(writer, sheet_name="行业统计")
            
            # 工作表4: 高分股票（A级以上）
            if "评级" in df.columns:
                high_rating = df[df["评级"].isin(["A+", "A"])]
                if not high_rating.empty:
                    high_rating.to_excel(writer, sheet_name="高分股票", index=False)
            
            # 工作表5: 高股息股票
            if "股息率_3年均值" in df.columns:
                high_dividend = df[df["股息率_3年均值"] >= 3].sort_values("股息率_3年均值", ascending=False)
                if not high_dividend.empty:
                    high_dividend.to_excel(writer, sheet_name="高股息股票", index=False)
            
            # 工作表6: 低估值股票
            if "PE_3年均值" in df.columns and "PB_3年均值" in df.columns:
                low_valuation = df[
                    (df["PE_3年均值"] < 15) & (df["PB_3年均值"] < 1.5)
                ].sort_values("PE_3年均值")
                if not low_valuation.empty:
                    low_valuation.to_excel(writer, sheet_name="低估值股票", index=False)
        
        logger.info(f"综合报告已导出: {filepath}")
        return filepath
    
    def generate_summary_report(self, df: pd.DataFrame) -> Dict:
        """
        生成汇总报告
        
        Args:
            df: 分析结果DataFrame
            
        Returns:
            汇总报告字典
        """
        report = {
            "生成时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "分析股票总数": len(df),
        }
        
        # 评级分布
        if "评级" in df.columns:
            report["评级分布"] = df["评级"].value_counts().to_dict()
        
        # 行业分布（前10）
        if "所属行业" in df.columns:
            report["行业分布(Top10)"] = df["所属行业"].value_counts().head(10).to_dict()
        
        # 关键指标统计
        key_indicators = ["PE_3年均值", "PB_3年均值", "股息率_3年均值", "价值评分"]
        for indicator in key_indicators:
            if indicator in df.columns:
                report[f"{indicator}_统计"] = {
                    "均值": df[indicator].mean(),
                    "中位数": df[indicator].median(),
                    "最小值": df[indicator].min(),
                    "最大值": df[indicator].max(),
                    "标准差": df[indicator].std()
                }
        
        # 高分股票数量
        if "评级" in df.columns:
            report["高分股票数量(A+和A)"] = len(df[df["评级"].isin(["A+", "A"])])
        
        # 高股息股票数量
        if "股息率_3年均值" in df.columns:
            report["高股息股票数量(>=3%)"] = len(df[df["股息率_3年均值"] >= 3])
        
        return report
    
    def save_summary_to_text(self, report: Dict, filename: str = None) -> str:
        """
        保存汇总报告为文本文件
        
        Args:
            report: 汇总报告字典
            filename: 文件名
            
        Returns:
            文件路径
        """
        if filename is None:
            filename = f"summary_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        filepath = os.path.join(self.output_dir, f"{filename}.txt")
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write("=" * 60 + "\n")
            f.write("A股基本面分析汇总报告\n")
            f.write("=" * 60 + "\n\n")
            
            for key, value in report.items():
                f.write(f"{key}:\n")
                if isinstance(value, dict):
                    for k, v in value.items():
                        if isinstance(v, float):
                            f.write(f"  {k}: {v:.2f}\n")
                        else:
                            f.write(f"  {k}: {v}\n")
                else:
                    f.write(f"  {value}\n")
                f.write("\n")
        
        logger.info(f"汇总报告已保存: {filepath}")
        return filepath


if __name__ == "__main__":
    # 测试代码
    exporter = DataExporter()
    
    # 创建测试数据
    test_df = pd.DataFrame({
        "股票代码": ["000001.SZ", "000002.SZ", "600000.SH"],
        "股票名称": ["平安银行", "万科A", "浦发银行"],
        "所属行业": ["银行", "房地产", "银行"],
        "PE_3年均值": [12.5, 15.2, 8.3],
        "PB_3年均值": [1.2, 1.5, 0.9],
        "股息率_3年均值": [3.5, 2.8, 4.2],
        "价值评分": [75, 65, 85],
        "评级": ["A", "B", "A+"]
    })
    
    print("测试导出功能...")
    
    # 测试Excel导出
    excel_path = exporter.export_to_excel(test_df, "test_export")
    print(f"Excel导出: {excel_path}")
    
    # 测试CSV导出
    csv_path = exporter.export_to_csv(test_df, "test_export")
    print(f"CSV导出: {csv_path}")
    
    # 测试综合报告
    report_path = exporter.export_comprehensive_report(test_df)
    print(f"综合报告: {report_path}")
    
    # 测试汇总报告
    summary = exporter.generate_summary_report(test_df)
    print("\n汇总报告:")
    for k, v in summary.items():
        print(f"{k}: {v}")
